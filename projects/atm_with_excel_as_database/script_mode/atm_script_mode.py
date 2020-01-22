"""

This script is a Skeleton for replicating ATM machine Basic Functionalites
Completing this script shall give a Strong Understanding of Python Programming

"""

from openpyxl import load_workbook

user_options  = {0 : 'exit',
                    1 : 'check_balance',
                    2 : 'withdraw_amount',
                    3 : 'deposit_amount'}


def choose_option():
    """
    displays all avaiable options to customer
    :return: user's input choice
    """

    print("*******************************************************")
    print("         Choose Option from below                ")
    print("*******************************************************")
    for i, j in user_options.items():
        print("{} ---> {}".format(i, j))

    option = input("Enter your choice :")
    return int(option)

def load_excel(read_only = True):
    """
    loads the excel sheet
    :return: active sheet object
    """
    wb = load_workbook('../bank_accounts.xlsx', read_only=read_only)
    sheet = wb.active
    return sheet

def change_cell_value(cell_position, new_value):
    """
    changes a value in the excel sheet cell
    :param cell_position: which cell needs to be changed
    :param new_value: new value to be placed
    :return: None
    """

    wb = load_workbook('../bank_accounts.xlsx')
    sheet = wb.active
    sheet[cell_position] = new_value
    wb.save('bank_accounts.xlsx')

def validate_user():
    """
    Validate the Username and Password using data from Excel Sheet
    :return:  True if Valid credentials , else False
    """
    input_username = input("Enter your user name :")
    input_password = input("Enter your password :")

    excel = load_excel()
    user_col = 'A'
    pass_col = 'C'
    row = 2
    while True:
        user_position = user_col + str(row)
        user = excel[user_position].value
        if user == input_username:
            print("valid Username...")
            password_position = pass_col + str(row)
            password = excel[password_position].value
            if password == input_password:
                print("Valid Password...")
                return True, row
            else:
                print("Wrong Password")
                return False, 0
        elif user == None:
            print("Invalid User...")
            return False, 0
        else:
            row += 1
            continue

def check_balance(user_position):
    """
    To check the balace amount of the user from Excel sheet
    :param user_position: row number of the user
    :return: balance amount of the user
    """
    bal_col = 'D'
    bal_position = bal_col + str(user_position)
    balance = load_excel()[bal_position].value
    return balance

def withdraw_amount(user_position):
    """
    To Withdraw amount - reduces the balance amount based on user input
    :param user_position: row number of the user
    :return: None
    """
    bal_col = 'D'
    user_withdraw_amount = int(input("Enter Amount to Withdraw :"))
    current_balance = check_balance(user_position)
    if current_balance <= user_withdraw_amount:
        print("InSufficient Balance...")
    else:
        cell = bal_col+str(user_position)
        new_amount = current_balance - user_withdraw_amount
        change_cell_value(cell, new_amount)
        print("Please collect cash...")




def deposit_amount(user_position):
    """
    To Deposit Money - Increase the current balance amount based on user deposited money
    :param user_position: row number of the user
    :return:
    """
    bal_col = 'D'
    cell = bal_col + str(user_position)
    user_deposit_amount = int(input("Enter Amount to Deposit :"))
    new_amount = check_balance(user_position) + user_deposit_amount
    change_cell_value(cell, new_amount)
    print("Your Amount is Deposited Successfully")



def atm_functionalities(user_position):
    """
    All Atm funcionalites are mapped based to user input choice
    :param user_position:
    :return: row number of the user
    """
    option = choose_option()

    if option == 1 :
        balance = check_balance(user_position)
        print("*************Your Balance Amount **************", balance)
        atm_functionalities(user_position)
    elif option == 2:
        withdraw_amount(user_position)
        atm_functionalities(user_position)
    elif option == 3:
        deposit_amount(user_position)
        atm_functionalities(user_position)
    elif option == 0:
        print("Thanks for Banking with us ...")
        exit()
    else:
        print("Invalid Input option")

def retry():
    """
    retry validate_user for 3 times and Exit if failed after 3 times
    :return:
    """
    print("Provided username password is Invalid... Tries left 2")
    count = 1
    while count <= 2:
        print("Retry Count :",count)
        validation, user_position = validate_user()
        if validation:
            atm_functionalities(user_position)
            break
        count += 1


def atm_run():
    """
    Starting point ...
    :return:
    """
    print("Hi Welcome !!!")
    validation, user_position = validate_user()
    if validation :
        atm_functionalities(user_position)
    else:
        retry()


if __name__ == "__main__":
    atm_run()
