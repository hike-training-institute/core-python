from openpyxl import load_workbook
from tkinter import END



def load_excel(read_only = True):
    """
    loads the excel sheet
    :return: active sheet object
    """
    wb = load_workbook('/home/nandagopal/hti/core-python/projects/atm_with_excel_as_database/bank_accounts.xlsx', read_only=read_only )
    sheet = wb.active
    return sheet

def change_cell_value(cell_position, new_value):
    """
    changes a value in the excel sheet cell
    :param cell_position: which cell needs to be changed
    :param new_value: new value to be placed
    :return: None
    """

    wb = load_workbook('bank_accounts.xlsx')
    sheet = wb.active
    sheet[cell_position] = new_value
    wb.save('bank_accounts.xlsx')

def validate_user(input_username, input_password):
    excel = load_excel()
    user_col = 'A'
    pass_col = 'C'
    row = 2
    while True:
        user_position = user_col + str(row)
        user = excel[user_position].value
        if user == input_username:
            password_position = pass_col + str(row)
            password = excel[password_position].value
            if password == input_password:
                return True, row
            else:
                return False, 0
        elif user == None:
            return False, 0
        else:
            row += 1
            continue


def check_balance(user_position, balance_entry=None):
    """
    To check the balace amount of the user from Excel sheet
    :param user_position: row number of the user
    :return: balance amount of the user
    """
    print("Checking Balance... ")

    bal_col = 'D'
    bal_position = bal_col + str(user_position)
    balance = load_excel()[bal_position].value
    if balance_entry:
        balance_entry.delete(0, END)
        balance_entry.insert(END, str(balance))
    return balance

def deposit_amount(user_deposit_amount, user_position):
    """
    To Deposit Money - Increase the current balance amount based on user deposited money
    :param user_position: row number of the user
    :return:
    """
    print("Depositing Amount ...")
    bal_col = 'D'
    cell = bal_col + str(user_position)
    new_amount = check_balance(user_position) + int(user_deposit_amount)
    change_cell_value(cell, new_amount)
    print("Your Amount is Deposited Successfully")


def withdraw_amount(user_withdraw_amount, user_position):
    """
    To Withdraw amount - reduces the balance amount based on user input
    :param user_position: row number of the user
    :return: None
    """
    print("Withdrawing Amount...")
    bal_col = 'D'
    current_balance = check_balance(user_position)
    if current_balance <= int(user_withdraw_amount):
        print("InSufficient Balance...")
    else:
        cell = bal_col+str(user_position)
        new_amount = current_balance - int(user_withdraw_amount)
        change_cell_value(cell, new_amount)
        print("Please collect cash...")